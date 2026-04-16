import pdfplumber
import re
import fitz
import glob
import pandas as pd
import numpy as np
import openpyxl
from copy import copy
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection
import os
import sys
import argparse
import unicodedata
from datetime import datetime

import random
import shutil
import defusedxml.ElementTree as ET
from typing import Tuple, List, Dict, Optional, Any
import core.config as config
from database import db_manager

def _es_token_color(token: Any) -> bool:
    if not token:
        return False
    normalizado = str(token).strip().upper().replace("-", "/")
    partes = [p.strip() for p in normalizado.split("/") if p.strip()]
    if not partes:
        return False

    if normalizado in config.COLORES_VALIDOS:
        return True

    return all(parte in config.COLORES_VALIDOS for parte in partes)


def extraer_datos_empresa(empresa_busqueda: str) -> Tuple[Optional[str], Optional[str]]:
    def _norm(texto: Any) -> str:
        txt = unicodedata.normalize("NFKD", str(texto or ""))
        txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
        txt = re.sub(r"[^a-zA-Z0-9]+", " ", txt).strip().lower()
        return txt

    def _puntaje_coincidencia(nombre_empresa: str, nombre_archivo: str) -> float:
        tokens_emp = [t for t in _norm(nombre_empresa).split() if len(t) > 2]
        if not tokens_emp:
            return 0
        archivo_norm = _norm(nombre_archivo)
        score = sum(1 for token in tokens_emp if token in archivo_norm)
        if _norm(nombre_empresa) in archivo_norm:
            score += 3
        return score

    # Buscar PDF CSF más parecido al nombre proporcionado
    archivos_pdf = glob.glob(os.path.join("machotes", "*CSF*.pdf"))
    if not archivos_pdf:
        print(f"No se encontró ningún archivo CSF en 'machotes' para '{empresa_busqueda}'.")
        return None, None

    ranked = sorted(
        archivos_pdf,
        key=lambda path: _puntaje_coincidencia(empresa_busqueda, os.path.basename(path)),
        reverse=True,
    )
    mejor_coincidencia = ranked[0]
    best_score = _puntaje_coincidencia(empresa_busqueda, os.path.basename(mejor_coincidencia))
    if best_score == 0:
        print(
            f"No hubo coincidencia directa para '{empresa_busqueda}'. "
            f"Usando CSF por defecto: {os.path.basename(mejor_coincidencia)}"
        )

    print(f"Extrayendo datos de CSF: {mejor_coincidencia}")
    doc = fitz.open(mejor_coincidencia)
    text = "\n".join(page.get_text() for page in doc[:2])
    doc.close()

    rfc = None
    razon_social = None

    # RFC flexible (persona moral/física)
    rfc_match = re.search(r"\b([A-Z&Ñ]{3,4}\d{6}[A-Z0-9]{3})\b", text, re.IGNORECASE)
    if rfc_match:
        rfc = rfc_match.group(1).upper()

    # Razón social con varias etiquetas posibles del CSF
    patrones_razon = [
        r"Denominaci[oó]n\s*/?\s*Raz[oó]n\s*Social\s*:?\s*(.+)",
        r"Denominaci[oó]n\s+o\s+Raz[oó]n\s+Social\s*:?\s*(.+)",
        r"Nombre,\s*denominaci[oó]n\s+o\s+raz[oó]n\s+social\s*:?\s*(.+)",
    ]
    for patron in patrones_razon:
        match = re.search(patron, text, re.IGNORECASE)
        if match:
            value = match.group(1).strip()
            if value and len(value) > 3:
                razon_social = value
                break

    # Fallback: a veces el valor viene en la siguiente línea
    if not razon_social:
        lines = [ln.strip() for ln in text.splitlines()]
        for idx, line in enumerate(lines):
            if re.search(r"Denominaci[oó]n|Raz[oó]n Social|raz[oó]n social", line, re.IGNORECASE):
                if idx + 1 < len(lines) and lines[idx + 1]:
                    razon_social = lines[idx + 1].strip()
                    break

    if not rfc or not razon_social:
        print(f"No se pudo extraer RFC o Razón Social con precisión del PDF {os.path.basename(mejor_coincidencia)}.")
    return rfc, razon_social


def obtener_empresas_csf() -> List[str]:
    archivos_pdf = glob.glob(os.path.join("machotes", "*CSF*.pdf"))
    empresas = []
    for path in archivos_pdf:
        nombre = os.path.basename(path)
        nombre = re.sub(r"\.pdf$", "", nombre, flags=re.IGNORECASE)
        nombre = re.sub(r"(?i)\bcsf\b", "", nombre)
        nombre = re.sub(r"[_\-]+", " ", nombre)
        nombre = re.sub(r"\s+", " ", nombre).strip()
        if nombre:
            empresas.append(nombre.upper())
    return sorted(set(empresas))



def load_data() -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if not db_manager.is_db_initialized():
        print("Base de datos SQLite no inicializada. Intentando migrar desde Excel...")
        if os.path.exists(config.PATH_INVENTARIO):
            if not db_manager.migrate_excel_to_sqlite(config.PATH_INVENTARIO):
                print("Aviso: No se pudo migrar Excel. Creando DB vacía.")
                db_manager.create_empty_inventory()
        else:
            print(f"No existe {config.PATH_INVENTARIO}. Inicializando BD limpia desde cero.")
            db_manager.create_empty_inventory()

    df_reporte, df_usados, df_xml = db_manager.get_inventory_dataframes()
        
    from core.state import AppState
    app_state = AppState()
    precio_col = app_state.config.get("columna_precio_default", "D1")

    if os.path.exists(config.PATH_PRECIOS):
        df_precios = pd.read_excel(config.PATH_PRECIOS, sheet_name='Para imprimir', header=2)
        cols_to_keep = ['CLAVE SAT', 'DESCRIPCION', 'MODELO']
        if precio_col in df_precios.columns:
            cols_to_keep.append(precio_col)
        elif 'D1' in df_precios.columns:
            cols_to_keep.append('D1')

        df_precios = df_precios[cols_to_keep]
        df_precios.dropna(subset=['MODELO'], inplace=True)
        df_precios['MODELO'] = df_precios['MODELO'].astype(str).str.strip().str.upper()
        df_precios['CLAVE SAT'] = df_precios['CLAVE SAT'].fillna(0).astype(int).astype(str)
    else:
        print(f"Advertencia: Archivo de precios '{config.PATH_PRECIOS}' no encontrado. Precios serán $0.")
        df_precios = pd.DataFrame(columns=['CLAVE SAT', 'DESCRIPCION', 'MODELO', precio_col])
    
    return df_reporte, df_usados, df_xml, df_precios

def aplicar_mapeo(modelo_base: str) -> str:
    if pd.isna(modelo_base):
        return modelo_base
    modelo_upper = str(modelo_base).strip().upper()
    return config.MAPEOS_MODELOS.get(modelo_upper, modelo_upper)

def procesar_inventario(
    df_reporte: pd.DataFrame,
    df_precios: pd.DataFrame,
    incluir_infantiles: bool = False,
    incluir_motobicis: bool = False,
    categoria: str = "TODAS",
    modelos: list = None,
    sucursales: list = None
) -> pd.DataFrame:
    df = df_reporte.copy()
    df['CLAVE SAT'] = df['CLAVE SAT'].astype(str)
    
    # 1. Filtro inicial optimizado
    if not incluir_motobicis:
        mask = ~df['DESCRIPCION'].astype(str).str.contains("MOTOCICLETA ELECTRICA", case=False, na=False)
        df = df[mask]
        
    # 2. Mapeo vectorizado usando replace y map
    # Creamos un diccionario de mapeo a partir de los datos únicos para evitar iterar uno por uno
    modelos_unicos = df['MODELO BASE'].dropna().unique()
    mapa_modelos = {m: aplicar_mapeo(m) for m in modelos_unicos}
    df['MODELO_MAPPED'] = df['MODELO BASE'].map(mapa_modelos).fillna(df['MODELO BASE'])
    
    from core.state import AppState
    app_state = AppState()
    precio_col = app_state.config.get("columna_precio_default", "D1")

    actual_precio_col = precio_col if precio_col in df_precios.columns else 'D1'
    if actual_precio_col not in df_precios.columns:
        df_precios[actual_precio_col] = pd.Series(dtype=float)

    # Preparar df_precios
    df_precios_unicos = df_precios.drop_duplicates(subset=['MODELO'], keep='last').set_index('MODELO')

    # 3. Join (merge) en lugar de iterar row por row
    # Hacemos un merge left para traer el precio, 'CLAVE SAT' y 'DESCRIPCION' desde df_precios
    df = df.merge(
        df_precios_unicos[[actual_precio_col, 'CLAVE SAT', 'DESCRIPCION']],
        left_on='MODELO_MAPPED',
        right_index=True,
        how='left',
        suffixes=('', '_precio')
    )

    # Actualizar valores si se encontraron en la tabla de precios
    precio_merge_col = f'{actual_precio_col}_precio' if f'{actual_precio_col}_precio' in df.columns else actual_precio_col
    mask_found = df[precio_merge_col].notna()
    df.loc[mask_found, precio_col] = pd.to_numeric(df.loc[mask_found, precio_merge_col], errors='coerce')
    if precio_col != 'D1':
        df['D1'] = df[precio_col] # Retrocompatibilidad
    df.loc[mask_found, 'CLAVE SAT'] = df.loc[mask_found, 'CLAVE SAT_precio'].astype(str)

    # Construir descripción vectorizada
    desc_base = df['DESCRIPCION_precio'].astype(str).str.strip().fillna('')
    modelo = df['MODELO_MAPPED'].astype(str).str.strip().fillna('')
    color = df['COLOR'].astype(str).str.strip()
    color = color.replace({'nan': '', 'None': ''})
    serie = df['No de SERIE:'].astype(str).str.strip().fillna('')

    # Vectorized string concatenation
    color_part = np.where(color != "", " " + color, "")
    desc_final = desc_base + " " + modelo + color_part + " NO DE SERIE:" + serie

    # Aplicar nueva descripción solo a los que cruzaron
    df.loc[mask_found, 'DESCRIPCION'] = desc_final[mask_found].str.upper()

    # Limpiar columnas temporales del merge
    df = df.drop(columns=[precio_merge_col, 'CLAVE SAT_precio', 'DESCRIPCION_precio'], errors='ignore')
            
    # 4. Filtros posteriores vectorizados
    if not incluir_infantiles:
        df = df[~df['DESCRIPCION'].astype(str).str.contains("INFANTIL ELECTRICO", case=False, na=False)]
        df = df[df['CLAVE SAT'].astype(str) != "60141000"]
        
    if sucursales and "TODAS" not in [str(s).upper() for s in sucursales]:
        sucursales_upper = [str(s).upper() for s in sucursales]
        df = df[df['SUCURSAL'].astype(str).str.upper().isin(sucursales_upper)]
        
    if categoria != "TODAS":
        pass
        
    if modelos and "TODOS" not in [str(m).upper() for m in modelos]:
        modelos_upper = [str(m).upper() for m in modelos]
        df = df[df['MODELO BASE'].astype(str).str.upper().isin(modelos_upper)]

    # 5. Limpieza final y cálculos matemáticos (vectorizados)
    df = df.dropna(subset=[precio_col])
    df = df[df[precio_col].astype(str).str.strip() != ""]
    df = df.dropna(subset=['CLAVE SAT'])
    
    df['CANTIDAD'] = 1
    df[precio_col] = pd.to_numeric(df[precio_col], errors='coerce')
    df['P. UNITARIO'] = df[precio_col] / 1.16
    df['SUBTOTAL'] = df['CANTIDAD'] * df['P. UNITARIO']
    df['IVA'] = df['SUBTOTAL'] * 0.16
    df['TOTAL'] = df['SUBTOTAL'] + df['IVA']

    return df

def seleccionar_articulos(df_disponibles: pd.DataFrame, monto_objetivo: float) -> pd.DataFrame:
    if df_disponibles.empty:
        return pd.DataFrame()

    # Extraemos solo los valores necesarios a arrays de numpy para iteración ultrarrápida
    # Preservamos los índices originales para reconstruir el DataFrame al final
    indices = df_disponibles.index.to_numpy()
    totales = df_disponibles['TOTAL'].to_numpy()
    num_items = len(totales)
    
    mejor_diferencia = float('inf')
    mejor_seleccion_indices = []
    
    # Reducimos iteraciones a 400 (suficiente para variedad dado que numpy es rápido
    # pero queremos evitar ciclos CPU excesivos). La aleatoriedad está garantizada.
    iteraciones = min(400, max(50, num_items * 2))

    for _ in range(iteraciones):
        # Barajamos un array de índices aleatorios
        orden = np.random.permutation(num_items)
        totales_shuffled = totales[orden]
        indices_shuffled = indices[orden]
        
        # Suma acumulada rápida
        cumsum = np.cumsum(totales_shuffled)

        # Encontramos dónde cortamos (Greedy phase)
        # Buscamos el último elemento donde la suma es <= monto_objetivo + 100
        corte_idx = np.searchsorted(cumsum, monto_objetivo + 100, side='right')

        if corte_idx == 0:
            continue

        seleccion_actual_idx = list(indices_shuffled[:corte_idx])
        suma_actual = cumsum[corte_idx - 1]

        diferencia_inicial = abs(monto_objetivo - suma_actual)

        # Fase 2: Local Search / Refinamiento Fino
        if diferencia_inicial > 0.01 and corte_idx < num_items:
            faltante = monto_objetivo - suma_actual

            # Buscar en los restantes
            sobrantes = totales_shuffled[corte_idx:]
            sobrantes_indices = indices_shuffled[corte_idx:]

            # Calculamos las diferencias de añadir cada pieza sobrante
            diferencias_relleno = np.abs(monto_objetivo - (suma_actual + sobrantes))

            # Encontramos la mejor mejora
            mejor_relleno_idx = np.argmin(diferencias_relleno)
            mejor_diff_relleno = diferencias_relleno[mejor_relleno_idx]

            if mejor_diff_relleno < diferencia_inicial:
                seleccion_actual_idx.append(sobrantes_indices[mejor_relleno_idx])
                suma_actual += sobrantes[mejor_relleno_idx]
                diferencia_final = mejor_diff_relleno
            else:
                diferencia_final = diferencia_inicial
        else:
            diferencia_final = diferencia_inicial

        # Evaluar la solución global
        if diferencia_final < mejor_diferencia:
            mejor_diferencia = diferencia_final
            mejor_seleccion_indices = seleccion_actual_idx
        elif abs(diferencia_final - mejor_diferencia) < 0.01 and len(seleccion_actual_idx) > len(mejor_seleccion_indices):
            mejor_seleccion_indices = seleccion_actual_idx
            
        if mejor_diferencia < 0.01:
            break
            
    df_resultado = df_disponibles.loc[mejor_seleccion_indices].copy()
    if not df_resultado.empty:
        df_resultado = df_resultado.sort_values(by=['SUCURSAL', 'MODELO BASE', 'No de SERIE:'])
        
    return df_resultado

def generar_machote(df_seleccion: pd.DataFrame, monto_objetivo: float, empresa: str, rfc: str, cuenta_mp: str) -> Tuple[str, str]:
    if not os.path.exists(config.OUTPUT_DIR):
        os.makedirs(config.OUTPUT_DIR)
        
    fecha_str = datetime.now().strftime("%d %b %Y").upper()
    total_real = df_seleccion['TOTAL'].sum()
    nombre_archivo = f"{fecha_str} ${total_real:,.2f} MACHOTE {empresa} {cuenta_mp}.xlsx"
    ruta_salida = os.path.join(config.OUTPUT_DIR, nombre_archivo)
    
    # Manejar caso donde el machote no exista
    if not os.path.exists(config.PATH_MACHOTE):
        print(f"ADVERTENCIA: No se encontró {config.PATH_MACHOTE}. Se generará uno vacío.")
        wb = openpyxl.Workbook()
        ws = wb.active
        # Basic headers just in case
        for i, header in enumerate(['CANTIDAD', 'UNIDAD', 'CLAVE SAT', 'DESCRIPCION', 'P. UNITARIO', 'SUBTOTAL', 'IVA', 'TOTAL']):
            ws.cell(row=7, column=i+2).value = header
    else:
        wb = openpyxl.load_workbook(config.PATH_MACHOTE)
        ws = wb.active
    
    for row in range(1, 10):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val == 'EMPRESA':
            ws.cell(row=row, column=3).value = empresa
        elif cell_val == 'RFC':
            ws.cell(row=row, column=3).value = rfc
            
    fila_inicio = 8
    
    estilos_cols = {}
    for col in range(2, 11):
        cell = ws.cell(row=8, column=col)
        estilos_cols[col] = {
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'number_format': cell.number_format,
            'alignment': copy(cell.alignment)
        }
        
    for r in range(8, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None
            
    col_idx = {col: i for i, col in enumerate(df_seleccion.columns)}
    for row in df_seleccion.itertuples(index=False, name=None):
        ws.cell(row=fila_inicio, column=2).value = row[col_idx['CANTIDAD']]
        ws.cell(row=fila_inicio, column=3).value = "PIEZA"
        ws.cell(row=fila_inicio, column=4).value = row[col_idx['CLAVE SAT']]
        ws.cell(row=fila_inicio, column=5).value = row[col_idx['DESCRIPCION']]
        ws.cell(row=fila_inicio, column=6).value = row[col_idx['P. UNITARIO']]
        ws.cell(row=fila_inicio, column=7).value = row[col_idx['SUBTOTAL']]
        ws.cell(row=fila_inicio, column=8).value = row[col_idx['IVA']]
        ws.cell(row=fila_inicio, column=9).value = row[col_idx['TOTAL']]
        
        for col in range(2, 11):
            cell = ws.cell(row=fila_inicio, column=col)
            estilo = estilos_cols.get(col)
            if estilo:
                cell.font = copy(estilo['font'])
                cell.border = copy(estilo['border'])
                cell.fill = copy(estilo['fill'])
                cell.number_format = estilo['number_format']
                cell.alignment = copy(estilo['alignment'])
        fila_inicio += 1
        
    fila_inicio += 1
    
    ws.cell(row=fila_inicio, column=5).value = "TOTALES"
    ws.cell(row=fila_inicio, column=5).font = Font(bold=True)
    ws.cell(row=fila_inicio, column=6).value = df_seleccion['P. UNITARIO'].sum()
    ws.cell(row=fila_inicio, column=7).value = df_seleccion['SUBTOTAL'].sum()
    ws.cell(row=fila_inicio, column=8).value = df_seleccion['IVA'].sum()
    ws.cell(row=fila_inicio, column=9).value = df_seleccion['TOTAL'].sum()
    
    for col in range(6, 10):
        cell = ws.cell(row=fila_inicio, column=col)
        cell.number_format = estilos_cols[col]['number_format'] if col in estilos_cols else '0.00'
        cell.font = Font(bold=True)
        cell.border = copy(estilos_cols[col]['border']) if col in estilos_cols else None
        
    wb.save(ruta_salida)
    return ruta_salida, nombre_archivo


def _guardar_warnings_pdf(ruta_pdf: str, warnings: List[str]) -> None:
    if not warnings:
        return
    os.makedirs(config.APP_DATA_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(config.PDF_WARNINGS_LOG, "a", encoding="utf-8") as fh:
        fh.write(f"\n[{timestamp}] {ruta_pdf}\n")
        for warning in warnings:
            fh.write(f"- {warning}\n")


def extraer_nuevos_articulos_excel(ruta_excel: str, with_report: bool = False) -> Any:

    articulos_encontrados = []
    warnings = []

    try:
        # Load the REPORTE sheet, assuming standard format with header at row 3 (0-indexed 3 is row 4)
        # We will try a few header rows just in case
        df = None
        for h in [3, 2, 1, 0]:
            try:
                temp_df = pd.read_excel(ruta_excel, sheet_name='REPORTE', header=h)
                if 'No de SERIE:' in temp_df.columns:
                    df = temp_df
                    break
            except Exception:
                pass

        if df is None:
            # Fallback to first sheet if REPORTE doesn't exist
            df = pd.read_excel(ruta_excel, header=0)
            if 'No de SERIE:' not in df.columns:
                # Try finding the series column heuristically
                for col in df.columns:
                    if 'SERIE' in str(col).upper():
                        df = df.rename(columns={col: 'No de SERIE:'})
                        break

        if 'No de SERIE:' not in df.columns:
            warnings.append(f"No se encontró la columna 'No de SERIE:' en el archivo {os.path.basename(ruta_excel)}.")
            if with_report:
                return [], warnings, {"error": "Columna de serie faltante"}
            return []

        df = df.dropna(subset=['No de SERIE:'])

        col_idx = {col: i for i, col in enumerate(df.columns)}
        for row in df.itertuples(index=False, name=None):
            serie_val = row[col_idx['No de SERIE:']] if 'No de SERIE:' in col_idx else ''
            serie = str(serie_val).strip()
            if len(serie) > 5 and serie.isalnum() and serie != "nan":
                sucursal_val = row[col_idx['SUCURSAL']] if 'SUCURSAL' in col_idx else 'ALMACEN'
                modelo_base_val = row[col_idx['MODELO BASE']] if 'MODELO BASE' in col_idx else (row[col_idx['MODELO']] if 'MODELO' in col_idx else 'SIN MODELO')
                color_val = row[col_idx['COLOR']] if 'COLOR' in col_idx else ''

                articulo = {
                    'SUCURSAL': str(sucursal_val).strip(),
                    'MODELO BASE': str(modelo_base_val).strip(),
                    'COLOR': str(color_val).strip().upper(),
                    'No de SERIE:': serie,
                    'CANTIDAD': 1
                }
                if articulo['COLOR'] == 'NAN': articulo['COLOR'] = ''
                if articulo['SUCURSAL'] == 'nan': articulo['SUCURSAL'] = 'ALMACEN'
                if articulo['MODELO BASE'] == 'nan': articulo['MODELO BASE'] = 'SIN MODELO'

                articulos_encontrados.append(articulo)

    except Exception as e:
        warnings.append(f"Error procesando Excel: {e}")

    report = {
        "lineas_analizadas": len(articulos_encontrados),
        "bloques_detectados": len(articulos_encontrados),
        "articulos_detectados": len(articulos_encontrados),
        "warnings": len(warnings),
        "warnings_log": config.PDF_WARNINGS_LOG,
    }

    if with_report:
        return articulos_encontrados, warnings, report
    return articulos_encontrados


def extraer_nuevos_articulos(ruta_pdf: str, with_report: bool = False) -> Any:
    
    match = re.search(r"reporte-productos (.+)\.pdf", os.path.basename(ruta_pdf), re.IGNORECASE)
    sucursal = match.group(1).strip() if match else "ALMACEN"
    
    articulos_encontrados = []
    seen_series = set()
    warnings = []
    bloques_detectados = 0
    
    with pdfplumber.open(ruta_pdf) as pdf:
        text = ""
        for p in pdf.pages:
            page_text = p.extract_text(layout=True) or ""
            if not page_text.strip():
                warnings.append(f"Página {p.page_number} sin texto extraíble.")
            text += page_text + "\n"
            
    lineas = text.split("\n")
    modelo_actual = ""
    color_actual = ""
    
    # Compile regex pattern outside loop for efficiency
    # Make the price part optional, capturing any text before it as series
    bloque_pattern = re.compile(r"^(.+?)\s+(\d+)\s+\(\d+\)\s+(.*?)(?:\s+\$[\d,]+\.\d{2})?$")

    for i in range(len(lineas)):
        linea = lineas[i].strip()
        if not linea or "Nombre" in linea or "Total de productos" in linea:
            continue
            
        match_bloque = bloque_pattern.match(linea)
        
        if match_bloque:
            bloques_detectados += 1
            nombre_completo = match_bloque.group(1).strip()
            partes_nombre = nombre_completo.split(" ")
            
            # Use combined color tokens for special compound cases like GRIS XUAN
            if len(partes_nombre) > 1:
                ultimo_token = partes_nombre[-1].upper()
                ultimos_dos_tokens = f"{partes_nombre[-2].upper()} {ultimo_token}" if len(partes_nombre) >= 2 else ultimo_token
            else:
                ultimo_token = ""
                ultimos_dos_tokens = ""

            if len(partes_nombre) >= 2 and _es_token_color(ultimos_dos_tokens):
                modelo_actual = " ".join(partes_nombre[:-2])
                color_actual = ultimos_dos_tokens.replace("-", "/")
            elif len(partes_nombre) >= 1 and _es_token_color(ultimo_token):
                modelo_actual = " ".join(partes_nombre[:-1])
                color_actual = ultimo_token.replace("-", "/")
            else:
                modelo_actual = nombre_completo
                color_actual = ""
                if i + 1 < len(lineas):
                    tokens = lineas[i + 1].strip().split()
                    if tokens:
                        siguiente = tokens[0]
                        if len(tokens) > 1:
                            siguientes_dos = f"{tokens[0]} {tokens[1]}"
                        else:
                            siguientes_dos = tokens[0]

                        if _es_token_color(siguientes_dos):
                             color_actual = siguientes_dos.upper().replace("-", "/")
                        elif _es_token_color(siguiente) and siguiente not in ["L1LTWT", "HWM7MTEZA", "LU5RMC"]:
                            color_actual = siguiente.upper().replace("-", "/")
                    else:
                        warnings.append(f"Línea {i + 2} vacía al intentar detectar color para '{nombre_completo}'.")
            
            series_str = match_bloque.group(3).strip()
            # Clean trailing commas
            if series_str.endswith(","):
                series_str = series_str[:-1]

            series = [s.strip() for s in series_str.split(',')]
            
            for s in series:
                # Ensure the string looks like a serial number and not just a price/time
                if len(s) > 5 and s.isalnum() and s not in seen_series:
                    articulos_encontrados.append({'SUCURSAL': sucursal, 'MODELO BASE': modelo_actual, 'COLOR': color_actual, 'No de SERIE:': s, 'CANTIDAD': 1})
                    seen_series.add(s)
            
            j = i + 1
            while j < len(lineas):
                lin_j = lineas[j].strip()

                # Stop immediately if this looks like a new block definition
                if bloque_pattern.match(lin_j):
                    break

                match_series_extra = re.findall(r"([A-Z0-9]{10,25})(?:,|$)", lin_j)
                if match_series_extra:
                    for s in match_series_extra:
                        if len(s) > 5 and s not in seen_series:
                            articulos_encontrados.append({'SUCURSAL': sucursal, 'MODELO BASE': modelo_actual, 'COLOR': color_actual, 'No de SERIE:': s, 'CANTIDAD': 1})
                            seen_series.add(s)

                # Stop if it has a price or time pattern indicating end of block
                if "$" in lin_j or re.match(r"^\d{2}:\d{2}:\d{2}$", lin_j):
                    break
                j += 1

    _guardar_warnings_pdf(ruta_pdf, warnings)
    report = {
        "lineas_analizadas": len(lineas),
        "bloques_detectados": bloques_detectados,
        "articulos_detectados": len(articulos_encontrados),
        "warnings": len(warnings),
        "warnings_log": config.PDF_WARNINGS_LOG,
    }
    if with_report:
        return articulos_encontrados, warnings, report
    return articulos_encontrados


def cargar_inventario(ruta_pdf: str, path_inventario: str, lista_articulos: Optional[List[Dict[str, Any]]] = None) -> str:
    
    print(f"Leyendo PDF de carga: {ruta_pdf}")
    if lista_articulos is not None:
        nuevos = lista_articulos
    else:
        nuevos = extraer_nuevos_articulos(ruta_pdf)
    print(f"Se encontraron {len(nuevos)} articulos en el PDF.")
    
    # Cargar lista de precios
    _, _, _, df_precios = load_data()
    df_precios_dict = df_precios.drop_duplicates(subset=['MODELO'], keep='last').set_index('MODELO').to_dict('index')
    
    wb = openpyxl.load_workbook(path_inventario)
    ws_reporte = wb['REPORTE']
    
    existentes = set()
    for sheet_name in ['REPORTE', 'USADOS', 'XML_ENCONTRADOS']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            col_serie = 6
            for c in range(1, 20):
                if ws.cell(row=4, column=c).value == 'No de SERIE:':
                    col_serie = c
                    break
            for r in range(5, ws.max_row + 1):
                val = ws.cell(row=r, column=col_serie).value
                if val: existentes.add(str(val).strip())
                
    a_insertar = []
    for art in nuevos:
        if art['No de SERIE:'] in existentes:
            print(f"ALERTA: Serie duplicada ignorada -> {art['No de SERIE:']}")
            continue
            
        # Calcular precios y descripciones cruzando con el catálogo
        modelo_busqueda = aplicar_mapeo(art['MODELO BASE'])
        
        d1_val = None
        clave_sat = None
        desc_base = ""
        
        from core.state import AppState
        app_state = AppState()
        precio_col = app_state.config.get("columna_precio_default", "D1")

        if modelo_busqueda in df_precios_dict:
            datos_precio = df_precios_dict[modelo_busqueda]
            try:
                d1_val = float(datos_precio.get(precio_col, datos_precio.get('D1')))
            except (ValueError, TypeError):
                d1_val = None
            clave_sat = str(datos_precio['CLAVE SAT'])
            desc_base = str(datos_precio['DESCRIPCION']).strip()
            
        color = art['COLOR'].strip()
        serie = art['No de SERIE:'].strip()
        
        if color:
            desc_final = f"{desc_base} {modelo_busqueda} {color} NO DE SERIE:{serie}"
        else:
            desc_final = f"{desc_base} {modelo_busqueda} NO DE SERIE:{serie}"
            
        art[precio_col] = d1_val
        if precio_col != 'D1':
            art['D1'] = d1_val
        art['CLAVE SAT'] = clave_sat
        art['DESCRIPCION'] = desc_final.upper()
        
        if pd.notna(d1_val):
            art['P. UNITARIO'] = d1_val / 1.16
            art['SUBTOTAL'] = art['P. UNITARIO'] * 1 # Cantidad 1
            art['IVA'] = art['SUBTOTAL'] * 0.16
            art['TOTAL'] = art['SUBTOTAL'] + art['IVA']
        else:
            art['P. UNITARIO'] = None
            art['SUBTOTAL'] = None
            art['IVA'] = None
            art['TOTAL'] = None
            print(f"ADVERTENCIA: Modelo {modelo_busqueda} no tiene precio {precio_col} en catálogo. Se cargará sin precio.")
            
        a_insertar.append(art)
            
    if not a_insertar:
        print("No hay articulos nuevos validos para insertar.")
        return
        
    print(f"Insertando {len(a_insertar)} articulos nuevos en SQLite...")
    # Solo actualizamos SQLite. Excel se genera bajo demanda después.
    db_manager.insert_new_items(a_insertar)
    
    path_salida = path_inventario # Ya no generamos _CARGADO temporalmente aquí
    print(f"Nuevos artículos insertados en base de datos. (Inventario Excel se actualizará bajo demanda).")
    return path_salida

def procesar_xmls(xml_dir: str) -> Dict[str, str]:
    
    archivos_xml = glob.glob(os.path.join(xml_dir, "*.xml"))
    if not archivos_xml:
        print(f"No se encontraron archivos XML en la carpeta {xml_dir}")
        return {}
        
    series_uuid = {}
    
    for archivo in archivos_xml:
        try:
            tree = ET.parse(archivo)
            root = tree.getroot()
            
            # Buscar el UUID (Esta en el Complemento TimbreFiscalDigital)
            uuid = None
            for elem in root.iter():
                # El tag suele ser {http://www.sat.gob.mx/TimbreFiscalDigital}TimbreFiscalDigital
                if 'UUID' in elem.attrib:
                    uuid = elem.attrib['UUID']
                    break
                    
            if not uuid:
                # Si el namespace es complejo, busqueda por expresion regular pura al archivo de texto
                with open(archivo, 'r', encoding='utf-8', errors='ignore') as f_xml:
                    contenido_xml = f_xml.read()
                    match_uuid = re.search(r'UUID="([a-fA-F0-9\-]+)"', contenido_xml)
                    if match_uuid:
                        uuid = match_uuid.group(1)
            
            if not uuid:
                print(f"No se encontró UUID en el archivo {archivo}")
                continue
                
            # Extraer todas las series del XML (Buscamos la palabra NO DE SERIE: o similar)
            # Analizando por expresion regular el contenido del archivo suele ser mas seguro y robusto en CFDI
            with open(archivo, 'r', encoding='utf-8', errors='ignore') as f_xml:
                contenido_xml = f_xml.read()
                
                # Buscar "NO DE SERIE:XXXXXX"
                matches_series = re.findall(r'NO DE SERIE:([A-Z0-9]+)', contenido_xml.upper())
                
                if matches_series:
                    for s in matches_series:
                        series_uuid[s.strip()] = uuid
                        
        except Exception as e:
            print(f"Error procesando el archivo XML {archivo}: {e}")
            
    print(f"Se encontraron {len(series_uuid)} series a buscar en los XMLs proporcionados.")
    return series_uuid
    
def actualizar_inventario_uuid(xml_dir: str, path_inventario: str) -> Tuple[List[str], str]:
    series_a_buscar = procesar_xmls(xml_dir)
    if not series_a_buscar:
        print("No hay UUIDs/series para actualizar.")
        return [], path_inventario
        
    print("Actualizando SQLite Inventario con UUIDs...")
    db_manager.mark_items_as_xml(series_a_buscar)
    print("Inventario SQLite actualizado con UUIDs.")
    return list(series_a_buscar.keys()), path_inventario

def actualizar_inventario_base(df_seleccion: pd.DataFrame, nombre_machote: str) -> str:
    print("Actualizando SQLite Inventario...")
    series_usadas = df_seleccion['No de SERIE:'].tolist()
    db_manager.mark_items_as_used(series_usadas, nombre_machote)
    return config.PATH_INVENTARIO

def importar_machote_externo(ruta_machote: str) -> Tuple[List[str], int]:

    print(f"Importando machote externo desde: {ruta_machote}")
    df_externo = pd.read_excel(ruta_machote, header=None)

    series_encontradas = []

    # Buscar posibles columnas de serie
    col_idx = {col: i for i, col in enumerate(df_externo.columns)}
    for row in df_externo.itertuples(index=False, name=None):
        # Vamos a buscar en todas las celdas el formato de serie
        for col in df_externo.columns:
            val = str(row[col_idx[col]]).strip()
            # Asumimos que una serie es algo como XXXXXX y típicamente alfanumérico

            # Buscar explícitamente "NO DE SERIE:"
            match = re.search(r"NO DE SERIE:\s*([A-Z0-9]+)", val, re.IGNORECASE)
            if match:
                s = match.group(1).strip()
                if len(s) > 5 and s not in series_encontradas:
                    series_encontradas.append(s)
            elif "SERIE" in val.upper():
                continue # Header row
            else:
                # Intento heurístico: Si es una celda sola y parece una serie (alfanumérica, >5 chars)
                # que está cerca de otras columnas descriptivas.
                pass

    if not series_encontradas:
        # Modo fallback: buscar en las columnas que parezcan series
        for col in df_externo.columns:
            for val in df_externo[col].dropna():
                val_str = str(val).strip()
                # Verificar si parece una serie pura alfanumerica
                if len(val_str) > 5 and val_str.isalnum():
                    if val_str not in series_encontradas:
                        series_encontradas.append(val_str)

    if not series_encontradas:
        print("No se encontraron números de serie válidos en el archivo externo.")
        return [], 0

    print(f"Series detectadas en el archivo externo: {len(series_encontradas)}")

    nombre_machote = os.path.basename(ruta_machote)
    series_a_marcar = set(series_encontradas)

    # Get available items to see which ones match
    df_rep, _, _ = db_manager.get_inventory_dataframes()

    if df_rep is None or df_rep.empty or "No de SERIE:" not in df_rep.columns:
        print("No hay inventario disponible para cruzar.")
        return series_encontradas, 0

    series_disponibles = set(df_rep["No de SERIE:"].astype(str).str.strip().tolist())

    series_coincidentes = series_a_marcar.intersection(series_disponibles)

    if series_coincidentes:
        db_manager.mark_items_as_used(list(series_coincidentes), f"EXT: {nombre_machote}")
        print("Machote externo importado exitosamente en SQLite.")

    return list(series_coincidentes), len(series_encontradas)


def deshacer_machote(nombre_machote: str) -> bool:

    print(f"Deshaciendo machote en SQLite: {nombre_machote}")

    conn = db_manager.get_connection()
    df_usados = pd.read_sql_query("SELECT * FROM inventario WHERE estado='USADO' AND machote=?", conn, params=(nombre_machote,))

    if df_usados.empty:
        print("No se encontraron artículos para este machote.")
        conn.close()
        return False

    series_a_restaurar = df_usados['no_serie'].tolist()

    try:
        cursor = conn.cursor()
        conn.execute("BEGIN TRANSACTION")
        placeholders = ','.join('?' * len(series_a_restaurar))
        cursor.execute(f'''
        UPDATE inventario
        SET estado = 'DISPONIBLE', machote = NULL
        WHERE no_serie IN ({placeholders}) AND estado = 'USADO'
        ''', series_a_restaurar)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Error revirtiendo machote en SQLite: {e}")
        raise e
    finally:
        conn.close()

    print("Machote deshecho exitosamente en SQLite.")
    return True

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--monto", type=float, required=False)
    parser.add_argument("--empresa", type=str, default="MOVILIDAD ELECTRICA DE JALISCO")
    parser.add_argument("--rfc", type=str, default="MEJ123456789")
    parser.add_argument("--cuenta", type=str, default="MP")
    parser.add_argument("--cargar", type=str, help="Ruta de un PDF de reporte de productos para ingresar al inventario")
    parser.add_argument("--xml_dir", type=str, help="Carpeta que contiene los archivos XML para validar UUIDs")
    args = parser.parse_args()
    
    if args.xml_dir:
        print(f"Modo Validacion XML activado para carpeta: {args.xml_dir}")
        actualizar_inventario_uuid(args.xml_dir, config.PATH_INVENTARIO)
        sys.exit(0)

    
    if args.cargar:
        print(f"Modo Carga activado para: {args.cargar}")
        cargar_inventario(args.cargar, config.PATH_INVENTARIO)
        sys.exit(0)

    # Intentar extraer datos del PDF de la Constancia de Situación Fiscal
    rfc_pdf, empresa_pdf = extraer_datos_empresa(args.empresa)
    
    rfc_final = rfc_pdf if rfc_pdf else args.rfc
    empresa_final = empresa_pdf if empresa_pdf else args.empresa
    
    print(f"Datos a utilizar para Machote -> Empresa: '{empresa_final}' | RFC: '{rfc_final}'")

    df_reporte, df_usados, df_xml, df_precios = load_data()
    df_disponibles = procesar_inventario(df_reporte, df_precios)
    df_seleccion = seleccionar_articulos(df_disponibles, args.monto)
    ruta_machote, nombre_machote = generar_machote(df_seleccion, args.monto, empresa_final, rfc_final, args.cuenta)
    
    actualizar_inventario_base(df_seleccion, nombre_machote)
    
if __name__ == '__main__':
    main()


def generar_machote_y_actualizar(df_seleccion: pd.DataFrame, monto_objetivo: float, empresa: str, rfc: str, cuenta_mp: str) -> Tuple[str, str, None]:
    ruta_machote, nombre_machote = generar_machote(df_seleccion, monto_objetivo, empresa, rfc, cuenta_mp)
    actualizar_inventario_base(df_seleccion, nombre_machote)
    # Devolvemos un 3er elemento None por compatibilidad con firmas antiguas
    return ruta_machote, nombre_machote, None


def cargar_inventario_y_reemplazar(ruta_pdf: str, path_inventario: str = config.PATH_INVENTARIO, lista_articulos: Optional[List[Dict[str, Any]]] = None) -> None:
    cargar_inventario(ruta_pdf, path_inventario, lista_articulos=lista_articulos)
    return None


def validar_xml_y_reemplazar(xml_dir: str, path_inventario: str = config.PATH_INVENTARIO) -> Tuple[None, List[str]]:
    series_actualizadas, _ = actualizar_inventario_uuid(xml_dir, path_inventario)
    return None, series_actualizadas
