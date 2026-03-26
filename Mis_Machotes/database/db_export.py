import pandas as pd
from database import db_manager
import os
from openpyxl.utils import get_column_letter

def export_inventory_to_excel(output_path):
    print(f"Exporting full database to {output_path}...")
    df_rep, df_us, df_xml = db_manager.get_inventory_dataframes()

    # We want to format the output similarly to the original Excel, but clean
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # Function to write and style a dataframe
        def write_styled_df(df, sheet_name):
            if df is None or df.empty:
                # write empty structure if no data
                df = pd.DataFrame(columns=["SUCURSAL", "MODELO", "MODELO BASE", "COLOR", "CANTIDAD", "No de SERIE:", "D1", "P. UNITARIO", "SUBTOTAL", "IVA", "TOTAL", "CLAVE SAT", "DESCRIPCION", "MACHOTE", "UUID"])

            # Drop unnecessary internal columns
            if 'FECHA ACTUALIZACION' in df.columns:
                df = df.drop(columns=['FECHA ACTUALIZACION'])
            if 'id' in df.columns:
                df = df.drop(columns=['id'])
            if 'estado' in df.columns:
                df = df.drop(columns=['estado'])

            # Rearrange columns slightly to look like original
            cols = ["SUCURSAL", "MODELO BASE", "MODELO", "COLOR", "CANTIDAD", "No de SERIE:", "D1", "P. UNITARIO", "SUBTOTAL", "IVA", "TOTAL", "CLAVE SAT", "DESCRIPCION"]
            if sheet_name == 'USADOS':
                cols.append("MACHOTE")
            if sheet_name == 'XML_ENCONTRADOS':
                cols.append("MACHOTE")
                cols.append("UUID")

            # Keep only columns that exist
            cols = [c for c in cols if c in df.columns]
            df = df[cols]

            df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Format totals row
            # Row 1 is empty, row 2 will have totals, row 3 empty, row 4 headers, row 5+ data
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4FAF6D")

            for col_idx, col_name in enumerate(cols, start=1):
                cell = worksheet.cell(row=4, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

                # Format money columns
                if col_name in ["D1", "P. UNITARIO", "SUBTOTAL", "IVA", "TOTAL"]:
                    # totals row formula
                    col_letter = get_column_letter(col_idx)
                    max_row = len(df) + 4
                    worksheet.cell(row=2, column=col_idx).value = f"=SUM({col_letter}5:{col_letter}{max_row})"
                    worksheet.cell(row=2, column=col_idx).font = Font(bold=True)
                    worksheet.cell(row=2, column=col_idx).number_format = '"$"#,##0.00'

                    # Data cells
                    for r in range(5, max_row + 1):
                        worksheet.cell(row=r, column=col_idx).number_format = '"$"#,##0.00'

        write_styled_df(df_rep, 'REPORTE')
        write_styled_df(df_us, 'USADOS')
        write_styled_df(df_xml, 'XML_ENCONTRADOS')

    print(f"Successfully generated clean inventory at {output_path}")
    return output_path
